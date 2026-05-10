"""XLSX workbook extractor (Week 2 multimodal expansion, Step 5).

Reads a multi-sheet Excel workbook of patient information using
openpyxl and produces a :class:`WorkbookXlsxFacts` model with
cell-coordinate citations. Like the docx extractor, this is text-only
— no Anthropic call — because the input has structural row/column
tagging that's better handled deterministically.

The cohort-5 workbook layout is four sheets — ``Patient`` (vertical
key/value), ``Medications`` (tabular), ``Labs_Trend`` (date-pivoted),
``Care_Gaps`` (tabular). The extractor's responsibility is to:

  1. Read the ``Patient`` sheet's column-A keys and pick out the
     fields we model (name, DOB, sex, MRN, etc.).
  2. Walk the ``Medications`` rows under the header row, emitting
     one :class:`WorkbookMedication` per data row.
  3. Unpivot ``Labs_Trend``: each cell at (test_row, date_col) becomes
     a :class:`WorkbookLabReading`, with the test name + LOINC + unit
     + ref range coming from the leftmost columns.
  4. Walk ``Care_Gaps`` rows under the header row.

Sheets that are missing or have an unexpected schema land their
section as the abstaining branch of the schema (``patient.name``
abstains with NO_DATA, lists return empty). The extractor does not
fail on missing optional sheets so a future workbook variant with
fewer sheets still partially extracts.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from anthropic import Anthropic
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from clinical_copilot.documents.schemas.citation import ExtractedField, SourceCitation
from clinical_copilot.documents.schemas.workbook_xlsx import (
    WorkbookCareGap,
    WorkbookLabReading,
    WorkbookMedication,
    WorkbookPatientInfo,
    WorkbookXlsxFacts,
)
from clinical_copilot.documents.synthetic_render import compute_line_bbox
from clinical_copilot.schemas.abstain import RuntimeAbstainReason

# Field-name → schema-attr mapping for the Patient sheet's vertical
# layout. Match keys are lowercased and have spaces / underscores
# stripped so a future "PCP NPI" or "pcp-npi" variant still maps.
_PATIENT_FIELD_MAP: dict[str, str] = {
    "name": "name",
    "dob": "dob",
    "sex": "sex",
    "mrn": "mrn",
    "pcp": "pcp",
    "pcpnpi": "pcp_npi",
    "phone": "phone",
    "address": "address",
    "insurance": "insurance",
    "allergies": "allergies",
}

_DATE_FIELDS: frozenset[str] = frozenset({"dob"})


def extract_workbook_xlsx(
    *,
    client: Anthropic,
    model: str,
    document_id: str,
    document_path: Path,
) -> WorkbookXlsxFacts:
    """Public registry entry point. Signature matches the other
    extractors so the registry dispatch is uniform."""

    del client, model  # text-only extractor

    # ``data_only=True`` makes openpyxl return the cached calculated
    # value of formula cells rather than the formula source. The
    # cohort-5 workbooks don't have formulas, but production exports
    # often do (e.g. spreadsheets that compute eGFR from creatinine).
    wb = load_workbook(str(document_path), data_only=True)

    patient = _extract_patient_sheet(document_id, wb["Patient"]) if "Patient" in wb.sheetnames else _empty_patient()
    medications = _extract_medications_sheet(document_id, wb["Medications"]) if "Medications" in wb.sheetnames else []
    lab_readings = _extract_labs_trend_sheet(document_id, wb["Labs_Trend"]) if "Labs_Trend" in wb.sheetnames else []
    care_gaps = _extract_care_gaps_sheet(document_id, wb["Care_Gaps"]) if "Care_Gaps" in wb.sheetnames else []

    return WorkbookXlsxFacts(
        document_id=document_id,
        patient=patient,
        medications=medications,
        lab_readings=lab_readings,
        care_gaps=care_gaps,
    )


# ---------------------------------------------------------------------
# Per-sheet extraction
# ---------------------------------------------------------------------


def _extract_patient_sheet(document_id: str, ws: Worksheet) -> WorkbookPatientInfo:
    """Vertical key/value: column A is the field name, column B is the
    value. Walk every row, look up the field name in the schema map,
    populate the corresponding field with a cell-coordinate citation.
    """

    fields: dict[str, ExtractedField[Any] | None] = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        if len(row) < 2:
            continue
        key_cell = row[0]
        value_cell = row[1]
        if key_cell.value is None or value_cell.value is None:
            continue
        normalized_key = _normalize_key(str(key_cell.value))
        attr = _PATIENT_FIELD_MAP.get(normalized_key)
        if attr is None:
            continue

        # WorkbookPatientInfo nests under WorkbookXlsxFacts.patient, so
        # every leaf path is prefixed with "patient." (e.g. "patient.name").
        cite = _cite(document_id, ws, value_cell, path=f"patient.{attr}")
        if attr in _DATE_FIELDS:
            parsed = _parse_date(value_cell.value)
            if parsed is None:
                fields[attr] = ExtractedField[date](
                    abstain_reason=RuntimeAbstainReason.OUT_OF_SCHEMA
                )
            else:
                fields[attr] = ExtractedField[date](value=parsed, citation=cite)
        else:
            value_str = str(value_cell.value).strip()
            if value_str == "":
                continue
            fields[attr] = ExtractedField[str](value=value_str, citation=cite)

    name_field = fields.get("name")
    if name_field is None:
        # Required; mark NO_DATA so downstream review surfaces the gap.
        name_field = ExtractedField[str](abstain_reason=RuntimeAbstainReason.NO_DATA)

    return WorkbookPatientInfo(
        name=name_field,
        dob=fields.get("dob"),
        sex=fields.get("sex"),
        mrn=fields.get("mrn"),
        pcp=fields.get("pcp"),
        pcp_npi=fields.get("pcp_npi"),
        phone=fields.get("phone"),
        address=fields.get("address"),
        insurance=fields.get("insurance"),
        allergies=fields.get("allergies"),
    )


def _extract_medications_sheet(document_id: str, ws: Worksheet) -> list[WorkbookMedication]:
    """Tabular layout. Header row → column index map; data rows below."""

    headers = _build_header_index(ws)
    out: list[WorkbookMedication] = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        # Skip rows where the generic name (the required column) is empty.
        generic_idx = headers.get("generic")
        if generic_idx is None:
            continue
        if generic_idx >= len(row) or row[generic_idx].value is None:
            continue
        generic_cell = row[generic_idx]
        generic_str = str(generic_cell.value).strip()
        if generic_str == "":
            continue

        med_index = len(out)
        med_prefix = f"medications[{med_index}]"
        out.append(
            WorkbookMedication(
                generic=ExtractedField[str](
                    value=generic_str,
                    citation=_cite(
                        document_id, ws, generic_cell, path=f"{med_prefix}.generic"
                    ),
                ),
                brand=_optional_str_from_row(
                    document_id, ws, row, headers, "brand", path=f"{med_prefix}.brand"
                ),
                strength=_optional_str_from_row(
                    document_id, ws, row, headers, "strength", path=f"{med_prefix}.strength"
                ),
                sig=_optional_str_from_row(
                    document_id, ws, row, headers, "sig", path=f"{med_prefix}.sig"
                ),
                indication=_optional_str_from_row(
                    document_id, ws, row, headers, "indication", path=f"{med_prefix}.indication"
                ),
                prescriber=_optional_str_from_row(
                    document_id, ws, row, headers, "prescriber", path=f"{med_prefix}.prescriber"
                ),
            )
        )

    return out


def _extract_labs_trend_sheet(document_id: str, ws: Worksheet) -> list[WorkbookLabReading]:
    """Date-pivoted layout: row 1 is headers (Test, LOINC, Units,
    Reference_Range, then one column per reading date). Each (test_row,
    date_col) cell becomes one ``WorkbookLabReading``.
    """

    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), None)
    if headers_row is None:
        return []

    # Identify the leftmost meta-data columns by name; remaining columns
    # are the trend dates.
    meta_indices: dict[str, int] = {}
    date_columns: list[tuple[int, date]] = []
    for idx, cell in enumerate(headers_row):
        raw = cell.value
        if raw is None:
            continue
        if isinstance(raw, str):
            normalized = _normalize_key(raw)
            if normalized in {"test", "loinc", "units", "referencerange"}:
                meta_indices[normalized] = idx
                continue
        # If not a known meta column, try parsing as a date — that's
        # what every column past the meta headers is.
        parsed = _parse_date(raw)
        if parsed is not None:
            date_columns.append((idx, parsed))

    test_idx = meta_indices.get("test")
    if test_idx is None or not date_columns:
        return []

    out: list[WorkbookLabReading] = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        if test_idx >= len(row) or row[test_idx].value is None:
            continue
        test_cell = row[test_idx]
        test_str = str(test_cell.value).strip()
        if test_str == "":
            continue

        # The leftmost meta cells (loinc/unit/reference_range) are shared
        # across every (test, date) reading we'll emit from this row, but
        # each reading is a separate WorkbookLabReading at its own index
        # in the lab_readings list — so we have to defer building those
        # ExtractedFields until we know the index, otherwise multiple
        # readings would alias the same path.
        for date_idx, reading_date in date_columns:
            if date_idx >= len(row):
                continue
            value_cell = row[date_idx]
            if value_cell.value is None:
                continue
            try:
                value_float = float(value_cell.value)
            except (TypeError, ValueError):
                continue

            reading_index = len(out)
            reading_prefix = f"lab_readings[{reading_index}]"
            out.append(
                WorkbookLabReading(
                    test=ExtractedField[str](
                        value=test_str,
                        citation=_cite(
                            document_id, ws, test_cell, path=f"{reading_prefix}.test"
                        ),
                    ),
                    value=ExtractedField[float](
                        value=value_float,
                        citation=_cite(
                            document_id, ws, value_cell, path=f"{reading_prefix}.value"
                        ),
                    ),
                    reading_date=ExtractedField[date](
                        value=reading_date,
                        citation=_cite(
                            document_id, ws, value_cell, path=f"{reading_prefix}.reading_date"
                        ),
                    ),
                    loinc=_optional_str_from_row(
                        document_id, ws, row, meta_indices, "loinc",
                        path=f"{reading_prefix}.loinc",
                    ),
                    unit=_optional_str_from_row(
                        document_id, ws, row, meta_indices, "units",
                        path=f"{reading_prefix}.unit",
                    ),
                    reference_range=_optional_str_from_row(
                        document_id, ws, row, meta_indices, "referencerange",
                        path=f"{reading_prefix}.reference_range",
                    ),
                )
            )

    return out


def _extract_care_gaps_sheet(document_id: str, ws: Worksheet) -> list[WorkbookCareGap]:
    headers = _build_header_index(ws)
    measure_idx = headers.get("measure")
    status_idx = headers.get("status")
    if measure_idx is None or status_idx is None:
        return []

    out: list[WorkbookCareGap] = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        if measure_idx >= len(row) or row[measure_idx].value is None:
            continue
        if status_idx >= len(row) or row[status_idx].value is None:
            continue

        measure_cell = row[measure_idx]
        status_cell = row[status_idx]
        measure_str = str(measure_cell.value).strip()
        status_str = str(status_cell.value).strip()
        if measure_str == "" or status_str == "":
            continue

        gap_index = len(out)
        gap_prefix = f"care_gaps[{gap_index}]"
        last_done_field = _optional_date_from_row(
            document_id, ws, row, headers, "lastdone", path=f"{gap_prefix}.last_done"
        )
        due_date_field = _optional_date_from_row(
            document_id, ws, row, headers, "duedate", path=f"{gap_prefix}.due_date"
        )
        notes_field = _optional_str_from_row(
            document_id, ws, row, headers, "notes", path=f"{gap_prefix}.notes"
        )

        out.append(
            WorkbookCareGap(
                measure=ExtractedField[str](
                    value=measure_str,
                    citation=_cite(
                        document_id, ws, measure_cell, path=f"{gap_prefix}.measure"
                    ),
                ),
                status=ExtractedField[str](
                    value=status_str,
                    citation=_cite(
                        document_id, ws, status_cell, path=f"{gap_prefix}.status"
                    ),
                ),
                last_done=last_done_field,
                due_date=due_date_field,
                notes=notes_field,
            )
        )

    return out


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------


def _empty_patient() -> WorkbookPatientInfo:
    return WorkbookPatientInfo(
        name=ExtractedField[str](abstain_reason=RuntimeAbstainReason.NO_DATA),
    )


def _normalize_key(text: str) -> str:
    """Lowercase, drop whitespace and underscores so 'PCP_NPI' / 'pcp npi'
    / 'PCPNPI' all collapse to the same lookup key."""

    return "".join(text.lower().split()).replace("_", "").replace("-", "")


def _build_header_index(ws: Worksheet) -> dict[str, int]:
    """Read row 1 as a header row and return a normalized-name → column-
    index map. Empty cells in the header row are skipped."""

    out: dict[str, int] = {}
    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), None)
    if headers_row is None:
        return out
    for idx, cell in enumerate(headers_row):
        if cell.value is None:
            continue
        out[_normalize_key(str(cell.value))] = idx
    return out


# Cache of (workbook_id, sheet_title) → list of 1-based source row
# numbers that the synthetic-text renderer keeps for that sheet
# (rows with at least one non-empty cell; matches the iteration in
# :func:`fetcher._render_xlsx`). Keyed off ``id(ws.parent)`` so a
# fresh workbook on a re-extraction doesn't get a stale entry.
_KEPT_ROWS_CACHE: dict[tuple[int, str], list[int]] = {}


def _kept_rows(ws: Worksheet) -> list[int]:
    """Return the 1-based source row numbers the renderer keeps for
    ``ws``. Cached per ``(workbook_id, sheet_title)`` so a workbook
    with a 5000-row sheet doesn't re-walk on every cite.
    """

    cache_key = (id(ws.parent), ws.title)
    cached = _KEPT_ROWS_CACHE.get(cache_key)
    if cached is not None:
        return cached
    rows: list[int] = []
    for row in ws.iter_rows():
        if any(cell.value not in (None, "") for cell in row):
            rows.append(row[0].row)
    _KEPT_ROWS_CACHE[cache_key] = rows
    return rows


def _cite(document_id: str, ws: Worksheet, cell: Cell, *, path: str) -> SourceCitation:
    """SourceCitation for a workbook cell, bound to a leaf path.

    ``page`` is the 1-indexed sheet number, matching the per-sheet
    page produced by the synthetic-text renderer. ``bbox`` is the
    normalized line band of the cell's row in that page, computed
    from the row's position among kept (non-empty) rows on the
    sheet. ``raw_text`` carries the ``Sheet!Cell`` reference for the
    review-page click-to-source UX. ``path`` is the JSON-pointer-
    style schema-walk position of the leaf this citation belongs to
    (e.g. ``"patient.name"``, ``"medications[2].sig"``) and is
    bound onto the citation's ``field_or_chunk_id``.

    Cells whose row is empty by the renderer's lights (no non-empty
    cells anywhere on the row, which can only happen when the
    extractor cites a programmatically-empty placeholder) get the
    full-page bbox as a safe fallback — the overlay still draws
    something rather than throwing.
    """

    sheet_index = ws.parent.sheetnames.index(ws.title) + 1
    kept = _kept_rows(ws)
    try:
        line_index = kept.index(cell.row)
    except ValueError:
        bbox: tuple[float, float, float, float] = (0.0, 0.0, 1.0, 1.0)
    else:
        bbox = compute_line_bbox(line_index, len(kept))
    return SourceCitation(
        document_id=document_id,
        page=sheet_index,
        bbox=bbox,
        confidence=1.0,
        raw_text=f"{ws.title}!{cell.coordinate}",
        field_or_chunk_id=path,
    )


def _optional_str_from_row(
    document_id: str,
    ws: Worksheet,
    row: tuple[Cell, ...],
    headers: dict[str, int],
    key: str,
    *,
    path: str,
) -> ExtractedField[str] | None:
    """``path`` is the schema-walk position of the leaf the returned
    field will be assigned to (e.g. ``"medications[0].brand"``)."""

    idx = headers.get(key)
    if idx is None or idx >= len(row):
        return None
    cell = row[idx]
    if cell.value is None:
        return None
    text = str(cell.value).strip()
    if text == "":
        return None
    return ExtractedField[str](
        value=text,
        citation=_cite(document_id, ws, cell, path=path),
    )


def _optional_date_from_row(
    document_id: str,
    ws: Worksheet,
    row: tuple[Cell, ...],
    headers: dict[str, int],
    key: str,
    *,
    path: str,
) -> ExtractedField[date] | None:
    """``path`` follows the same convention as :func:`_optional_str_from_row`."""

    idx = headers.get(key)
    if idx is None or idx >= len(row):
        return None
    cell = row[idx]
    if cell.value is None:
        return None
    parsed = _parse_date(cell.value)
    if parsed is None:
        return ExtractedField[date](abstain_reason=RuntimeAbstainReason.OUT_OF_SCHEMA)
    return ExtractedField[date](
        value=parsed,
        citation=_cite(document_id, ws, cell, path=path),
    )


def _parse_date(raw: object) -> date | None:
    """Best-effort date parse. openpyxl returns ``datetime.date`` /
    ``datetime.datetime`` natively for Excel date cells, but some
    workbooks store dates as ISO strings — handle both."""

    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(raw.strip(), fmt).date()
            except ValueError:
                continue
    return None
